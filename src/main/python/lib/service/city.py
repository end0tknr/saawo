#!python
# -*- coding: utf-8 -*-

from psycopg2  import extras # for bulk insert
import appbase
import openpyxl # for xlsx
import xlrd     # for xls
import os
import re
import sys
import tempfile
import time
import urllib.request

# https://www.soumu.go.jp/denshijiti/code.html

master_src_url = "https://www.soumu.go.jp/main_content/000925835.xls"
#master_src_url = "https://www.soumu.go.jp/main_content/000730858.xlsx"
master_excel = "000925835.xls"
bulk_insert_size = 20

seirei_city_names = (
    "大阪市","名古屋市","京都市","横浜市","神戸市","北九州市",
    "札幌市","川崎市","福岡市","広島市","仙台市","千葉市",
    "さいたま市","静岡市","堺市","新潟市","浜松市","岡山市",
    "相模原市","熊本市")

re_exp_pref = "神奈川県|和歌山県|鹿児島県|.{2}[都道府県]"
re_exp_city = "|".join([
    "東村山市","武蔵村山市","羽村市","四日市市","廿日市市","大村市","野々市市",
    "玉村町","大町町","大町市","田村市","十日町市","八丈町"] )
    
re_zipcode = re.compile("〒\d\d\d-\d\d\d\d")
re_pref_cities = [re.compile("^(%s).*?(%s)(.*)"  % (re_exp_pref,re_exp_city) ),
                  re.compile("^(%s)(.+?市.+?区)(.*)"    % (re_exp_pref) ),
                  re.compile("^(%s)(.+?[市区町村])(.*)" % (re_exp_pref) ),
                  re.compile("^(%s).+?郡(.+?[町村])(.*)"% (re_exp_pref) ) ]
logger = appbase.AppBase().get_logger()

class CityService(appbase.AppBase):
    
    def __init__(self):
        pass


    def calc_save_lnglat(self):
        func_name = sys._getframe().f_code.co_name
        logger.info("start %s",func_name)
        
        browser = self.get_browser()
        re_compile = re.compile("/@(\d+.\d+),(\d+\.\d+)\,\d+z/")

        max_retry = 10
        
        cities = self.get_all()
        i = 0
        for city in cities:
            i+=1
            if i % 20 == 0:
                logger.info("%s %s %s", func_name,city["pref"],city["city"])
                
            if city["lng"] and city["lat"]:
                continue
            
            
            pref_city = city["pref"] + city["city"]
            req_url = "https://www.google.com/maps/place/" + pref_city
            browser.get(req_url)
            i = 0
            longitude = None
            latitude  = None
            while i < max_retry:
                i += 1
                time.sleep(1)
                re_result = re_compile.search( browser.current_url )
                if not re_result:
                    continue
                longitude = re_result.group(1) # 緯度
                latitude  = re_result.group(2) # 経度
                self.save_lnglat(city, longitude, latitude )
                break

            if not longitude or not latitude:
                logger.error("fail calc lng lat for %s" % (pref_city,))

        browser.close()

    def save_lnglat(self,city,longitude,latitude):
        logger.info("%s %s %s %s"%(city["pref"],city["city"],longitude,latitude))
        
        sql = """
UPDATE city set lng=%s, lat=%s where code=%s 
"""
        sql_args = (longitude,latitude,city["code"])

        db_conn = self.db_connect()
        with self.db_cursor(db_conn) as db_cur:
            try:
                db_cur.execute(sql,sql_args)
                db_conn.commit()
            except Exception as e:
                    logger.error(e)
                    logger.error(sql)
                    logger.error(sql_args)
                    return False
        return True
            
    def get_near_cities(self, pref,city):
        sql = """
SELECT c.*
FROM city c
JOIN near_city nc
  ON (c.pref=nc.near_pref AND c.city=nc.near_city)
WHERE nc.pref=%s AND nc.city=%s
"""
        sql_args = (pref,city)
        ret_datas = []
        db_conn = self.db_connect()
        with self.db_cursor(db_conn) as db_cur:
            try:
                db_cur.execute(sql,sql_args)
            except Exception as e:
                logger.error(e)
                logger.error(sql)
                return []
            
            for ret_row in  db_cur.fetchall():
                ret_datas.append( dict( ret_row ))
            
        return ret_datas
        
        
    def save_near_cities(self, pref,city, near_cities):
        sql = """
INSERT INTO near_city (pref,city,near_pref,near_city) VALUES (%s,%s,%s,%s)
"""
        db_conn = self.db_connect()
        with self.db_cursor(db_conn) as db_cur:
            for near_city in near_cities:
                sql_args = (pref,city, near_city["pref"], near_city["city"])
                #print(sql_args)
                try:
                    db_cur.execute(sql,sql_args)
                except Exception as e:
                    logger.error(e)
                    logger.error(sql)
                    return False
            db_conn.commit()
        return True
            
        
    def get_all(self):
        ret_data = []
        sql = """
SELECT *
FROM city
"""
        db_conn = self.db_connect()
        with self.db_cursor(db_conn) as db_cur:
            try:
                db_cur.execute(sql)
            except Exception as e:
                logger.error(e)
                logger.error(sql)
                return []
            
            for ret_row in  db_cur.fetchall():
                ret_data.append( dict( ret_row ))
        return ret_data
    
    def get_all_pref_city(self):
        ret_data = []
        sql = """
SELECT pref,city
FROM city
GROUP BY pref,city
ORDER BY pref,city
"""

        db_conn = self.db_connect()
        with self.db_cursor(db_conn) as db_cur:
            try:
                db_cur.execute(sql)
            except Exception as e:
                logger.error(e)
                logger.error(sql)
                return []
            
            for ret_row in  db_cur.fetchall():
                ret_data.append( dict( ret_row ))

        return ret_data

    def get_seirei_wards(self, city_name):
        ret_data = []
        sql = "select * from city where city like '%s'"
        city_name = city_name + "%区"

        db_conn = self.db_connect()
        with self.db_cursor(db_conn) as db_cur:
            try:
                db_cur.execute(sql % (city_name) )
            except Exception as e:
                logger.error(e)
                logger.error(sql)
                return []
            
            for ret_row in  db_cur.fetchall():
                ret_data.append( dict( ret_row ))

        return ret_data

        
    def get_seirei_cities(self):
        ret_data = []
        sql = "select * from city where city in %s"
        
        with self.db_connect() as db_conn:
            with self.db_cursor(db_conn) as db_cur:
                try:
                    db_cur.execute(sql  % str(seirei_city_names) )
                    for ret_row in  db_cur.fetchall():
                        ret_data.append( dict( ret_row ))
                except Exception as e:
                    logger.error(e)
                    logger.error(sql)
                    return []
        return ret_data
    
    def is_seirei_city(self, city_name):
        if not city_name:
            return False
        
        for seirei_city_name in seirei_city_names:
            if seirei_city_name in city_name:
                return True
        return False
        
    def download_master(self):
        func_name = sys._getframe().f_code.co_name

        re_pat = re.compile("\.(xlsx|xls)$",re.IGNORECASE)
        re_result = re_pat.search(master_src_url)
        if not re_result:
            logger.error("bad src url %s", master_src_url)
            return []

        ext = re_result.group(1).lower()
        if ext=="xlsx":
            return self.download_master_xlsx(master_src_url)
        if ext=="xls":
            return self.download_master_xls(master_src_url)

        logger.error("bad src url %s", master_src_url)
        return []
        
    def download_master_xls(self,master_src_url):
        func_name = sys._getframe().f_code.co_name
        logger.info("start %s %s",func_name,master_src_url)
        
        ret_data = []
        
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_excel_path =os.path.join(tmp_dir, master_excel)
        
            try:
                data = urllib.request.urlopen(master_src_url).read()

                with open(tmp_excel_path, mode="wb") as fh:
                    fh.write(data)

                wbook = xlrd.open_workbook(tmp_excel_path)
                for sheetname in wbook.sheet_names():
                    wsheet = wbook.sheet_by_name(sheetname)
                    logger.info("start %s %d rows", sheetname, wsheet.nrows)

                    tmp_ret_data = self.__load_xls_wsheet( wsheet )
                    ret_data.extend( tmp_ret_data )


            except Exception as e:
                logger.error("fail", master_src_url)
                logger.error(e)
                return []
            
            return ret_data
    

    def download_master_xlsx(self,master_src_url):
        func_name = sys._getframe().f_code.co_name
        logger.info("start %s %s",func_name,master_src_url)
        
        ret_data = []
        
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_xlsx_path =os.path.join(tmp_dir, master_xlsx)
        
            try:
                data = urllib.request.urlopen(master_src_url).read()

                with open(tmp_xlsx_path, mode="wb") as fh:
                    fh.write(data)

                wbook = openpyxl.load_workbook(tmp_xlsx_path)
                for sheetname in wbook.sheetnames:
                    logger.info("start %s %d rows" %
                                (sheetname, wbook[sheetname].max_row) )

                    tmp_ret_data = self.__load_xlsx_wsheet( wbook[sheetname] )
                    ret_data.extend( tmp_ret_data )


            except Exception as e:
                logger.error("fail", master_src_url)
                logger.error(e)
                return []
            
            return ret_data
    

        
    def parse_pref_city(self, address_org):

        # refer to https://ja.wikipedia.org/wiki/%E9%A0%88%E6%81%B5%E7%94%BA
        conv_names = [
            ("須惠町","須恵町"),("茅ケ崎","茅ヶ崎"),("駒ケ根","駒ヶ根"),
            ("鶴ケ島","鶴ヶ島"),("七ケ浜","七ヶ浜"),("七ケ宿","七ヶ宿"),
            ("五ケ瀬","五ヶ瀬"),("檮原町","梼原町")
        ]
        for conv_name in conv_names:
            address_org = address_org.replace(conv_name[0],conv_name[1])

        address_org = re_zipcode.sub("",address_org)
        
        re_compile_space = re.compile("[\s\n]*")
        address_org = re_compile_space.sub("", address_org )
        
        for re_compile in re_pref_cities:
            re_result = re_compile.search(address_org)
            
            if not re_result:
                continue
            
            pref  = re_result.group(1)
            city  = re_result.group(2)
            other = re_result.group(3)

            if self.find_def_by_pref_city(pref, city):
                return [pref,city,other]
        
        logger.warning("fail parse "+address_org)
        return ["?","?",address_org]
        
    def  __load_xls_wsheet(self, wsheet):
        ret_data = []

        headers = []
        col_no = 0
        # load headers
        while col_no < wsheet.ncols :
            header = wsheet.cell(0,col_no).value
            if not header:
                col_no += 1
                continue
                
            headers.append( header.replace("\n","") )
            col_no += 1

        logger.debug(headers)

        row_no = 1
        while row_no < wsheet.nrows :
            col_no = 0
            ret_row = {}
            for header in headers:
                tmp_val =  wsheet.cell(row_no, col_no).value
                if header=="団体コード":
                    tmp_val = int(tmp_val)
                ret_row[header] = tmp_val
                col_no += 1

            if not ret_row["都道府県名（漢字）"] and \
               not ret_row["市区町村名（漢字）"] :
                logger.info("break at %d/%d row" % (row_no,wsheet.max_row) )
                break

            ret_data.append(ret_row)
            row_no += 1
            
        return ret_data

    def  __load_xlsx_wsheet(self, wsheet):
        ret_data = []

        headers = []
        col_no = 1
        # load headers
        while col_no < wsheet.max_column :
            header = wsheet.cell(column=col_no, row=1).value
            if not header:
                col_no += 1
                continue
                
            headers.append( header.replace("\n","") )
            col_no += 1
            
        logger.debug(headers)

        row_no = 2
        while row_no < wsheet.max_row :
            col_no = 1
            ret_row = {}
            for header in headers:
                ret_row[header] = wsheet.cell(column=col_no, row=row_no).value
                col_no += 1

            if not ret_row["都道府県名（漢字）"] and \
               not ret_row["市区町村名（漢字）"] :
                logger.info("break at %d/%d row" % (row_no,wsheet.max_row) )
                break

            ret_data.append(ret_row)
            row_no += 1
            
        return ret_data

    def del_all_tbl_rows(self):
        logger.info("start")
        
        with self.db_connect() as db_conn:
            with self.db_cursor(db_conn) as db_cur:
                sql = "DELETE FROM city"
                try:
                    db_cur.execute(sql)
                    db_conn.commit()
                except Exception as e:
                    logger.error(e)
                    return False
        return True
    

    def find_def_by_code_city(self,code,city):
        sql = "SELECT * from city where code like %s and city like %s"
        sql_args = ("%"+code+"%", "%"+ city,)
        
        with self.db_connect() as db_conn:
            with self.db_cursor(db_conn) as db_cur:
                try:
                    db_cur.execute(sql, sql_args)
                except Exception as e:
                    logger.error(e)
                    logger.error(city)
                    
                    return {}

                ret_rows = db_cur.fetchall()
                if len(ret_rows) == 0:
                    return None
                
                return dict( ret_rows[0] )

            
    def find_def_by_pref_city(self,pref, city):
        sql = "SELECT * from city where pref = %s and city = %s"
        sql_args = (pref,city)
        
        with self.db_connect() as db_conn:
            with self.db_cursor(db_conn) as db_cur:
                try:
                    db_cur.execute(sql, sql_args)
                except Exception as e:
                    logger.error(e)
                    logger.error(city)
                    return {}

                ret_rows = db_cur.fetchall()
                if len(ret_rows) == 0:
                    return None
                
                return dict( ret_rows[0] )
                
    
    def find_defs_by_city(self,city):
        sql = "SELECT * from city where city = %s"
        sql_args = (city,)
        
        ret_datas = []
        db_conn = self.db_connect()
        
        with self.db_cursor(db_conn) as db_cur:
            try:
                db_cur.execute(sql, sql_args)
            except Exception as e:
                logger.error(e)
                logger.error(city)
                return []

            for ret_row in  db_cur.fetchall():
                ret_datas.append( dict( ret_row ) )
                
        return ret_datas
    
    def save_tbl_rows(self, rows):
        logger.info("start")
        
        row_groups = self.__divide_rows(rows, bulk_insert_size)
        sql = \
            "INSERT INTO city(code,pref,city) values %s ON CONFLICT DO NOTHING"
        
        with self.db_connect() as db_conn:
            with self.db_cursor(db_conn) as db_cur:

                for row_group in row_groups:
                    try:
                        # bulk insert
                        extras.execute_values(db_cur,sql,row_group)
                    except Exception as e:
                        logger.error(e)
                        logger.error(sql)
                        logger.error(row_group)
                        return False
                    
            db_conn.commit()
        return True

    def __divide_rows(self, org_rows, chunk_size):
        
        i = 0
        chunk = []
        ret_rows = []
        for org_row in org_rows:
            
            chunk.append( ( org_row['団体コード'],
                            org_row['都道府県名（漢字）'],
                            org_row['市区町村名（漢字）'] or"" ))
            
            if len(chunk) >= chunk_size:
                ret_rows.append(chunk)
                chunk = []
            i += 1

        if len(chunk) > 0:
            ret_rows.append(chunk)

        return ret_rows
